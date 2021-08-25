from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

import os

pw = '/Users/files/work/cooperate/udn/cases/3_UDN730501'
os.chdir(pw)


dest_filename = 'test.xlsx'

# create workbook
wb = Workbook()


# side style
double = Side(border_style="double", color="ff0000")
thin = Side(border_style="thin", color="000000")

border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="DDDDDD")
fill_grad = GradientFill(stop=("000000", "FFFFFF"))
ft_bold = Font(b=True, color="FF0000")
align = Alignment(horizontal="center", vertical="center")



# swich sheet
ws1 = wb.active
# sheet rename
ws1.title = "Original"

ws1.merge_cells('A1:O1')
title = ws1['A1']
title.value = 'UDN12345'
title.font = ft_bold


wb.save(dest_filename)
