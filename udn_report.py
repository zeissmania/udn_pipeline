"""

"""
import sys, os, re
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle, Color
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

import xlsxwriter


# # def s():
# #     fno = 'test.xlsx'
# #     wb.save(fno)
# #     !open {fno}
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = 'Original'


def main(prj, pw=None, fn_selected_genes=None, sv_caller='dragen'):
    """
    expected file = selected.genes.xlsx
    prj = UDNxxxx
    pw, default is current folder
    fn_selected_genes , default is '{prj}.selected.genes.txt'
    first column should be the sv_type, valid = 'denovo, heter, xlink'
    other columns are just as the final tsv file generated by the pipeline

    # family info, is a list or tuple, like
    # ['Proand 123456', 'Mother (unaff) 22222', 'Father(unaff) 33333']
    """


    pw = pw or os.getcwd()
    for fn_tmp in [fn_selected_genes, f'{pw}/{prj}.selected.genes.xlsx', f'{pw}/selected.genes.xlsx']:
        if fn_tmp and os.path.exists(fn_tmp):
            fn_selected_genes = fn_tmp
            break
    else:
        print(f'ERROR, selected.genes.xlsx file not found!')
        return 1

    # group the selected SV into groups
    sv_selected = {}
    sv_type_conversion = {'denovo': 'De Novo',
                        'de novo': 'De Novo',
                        'd': 'De Novo',
                        'de': 'De Novo',
                        'heter': 'Heterozygous',
                        'he': 'Heterozygous',
                        'h': 'Heterozygous',
                        'hetero': 'Heterozygous',
                        'het': 'Heterozygous',
                        'xlink': 'X-linked',
                        'x-link': 'X-linked',
                        'x': 'X-linked',
                        'xl': 'X-linked',
                        }


    wb = xlsxwriter.Workbook(f'{pw}/{prj}.report.xlsx')
    ws = wb.add_worksheet('Original')


    # read the raw data
    d = pd.read_excel(fn_selected_genes, keep_default_na=False)
    header = list(d.columns)
    idx_copy_nmber = header.index('annot_ranking') + 1


    prev_columns = ["mut_type", "AMELIE", 'filter', "anno_id", "chr_", "pos_s", "pos_e", "sv_type", "qual", "exon_span_tag", "gn", "sv_len", "exon_span", "gain_af_max", "loss_af_max", 'gain_source', 'loss_source', "ddd_mode", "ddd_disease", "omim", "phenotype", "inheritance", "annot_ranking"]

    # n -1 because, in the following info list, the first element , mut_type is removed
    idx = {_: n - 1 for n, _ in enumerate(prev_columns)}


    idx_col_rank =header.index('AMELIE')

    # sort the rank
    col_rank = d.columns[idx_col_rank]
    d_rank = {}
    for gn, score in zip(d['gn'], d[col_rank]):
        try:
            d_rank[gn] = min(d_rank[gn], score)
        except:
            d_rank[gn] = score

    d_rank = sorted(d_rank.items(), key=lambda _: _[1])
    d_rank = {k[0]: n + 1 for n, k in enumerate(d_rank)}
    d[col_rank] = d['gn'].map(d_rank)

    # sort the family
    family_info_raw = list(header[idx_copy_nmber:])
    family_info_priority = []
    d_family_priority = {'proband': 1, 'mother': 2, 'father':3, 'sister': 4, 'brother': 5}

    for i in family_info_raw:
        rel_to_proband = i.split(' ', 1)[0].lower()
        rel_to_proband = re.split('\W+', rel_to_proband)[0]
        try:
            family_info_priority.append(d_family_priority[rel_to_proband])
        except:
            family_info_priority.append(999)

    family_info = sorted(zip(family_info_raw, family_info_priority), key=lambda x: x[1])
    family_info = [_[0] for _ in family_info]

    header = prev_columns + family_info
    d = d[header]

    # data = [_.strip().split('\t') for _ in fh]

    bad = 0
    for i in d.itertuples(False, None):
        sv_type = i[0]
        try:
            sv_type.lower()
        except:
            print(i)
            print(d.head())
            sys.exit(1)

        if sv_type.lower() not in sv_type_conversion:
            if not sv_type.strip():
                print(f'mut type not specified: {i}')
                bad = 1
            try:
                if i[1] != "AMELIE":
                    print(f'wrong SV catergory found! svtype={sv_type}')
            except:
                print(f'wrong line: {i}')
            continue
        sv_type_long = sv_type_conversion[sv_type.lower()]
        try:
            sv_selected[sv_type_long].append(i[1:])
        except:
            sv_selected[sv_type_long] = [i[1:]]

    if bad:
        print('Fail to build report, add mut type of each SV in selected.genes first')
        return 1
    n_family = len(family_info)
    total_col = 11 + n_family


    # define formats
    formats = {}
    header1 = wb.add_format({'font_size': 16, 'bold': True, })
    header2 = wb.add_format({'font_size': 14, 'bold': True, })
    formats['header_sv'] = wb.add_format({'font_size': 14, 'bold': True, 'bg_color': '#d9d9d9', 'text_wrap': True})
    formats['header_main'] = wb.add_format({'font_size': 11, 'bold': True, 'bg_color': '#dce6f1', 'border': 2, 'border_color': '#d9d9d9', 'text_wrap': True})
    formats['header_main'].set_align('center')
    formats['header_main'].set_align('vcenter')

    formats['fmt_txt'] = wb.add_format({'font_size': 11, 'bold': False, 'border': 2, 'border_color': '#d9d9d9'})
    formats['fmt_txt'].set_align('center')
    formats['fmt_txt'].set_align('vcenter')


    formats['fmt_wrap'] = wb.add_format({'font_size': 11, 'bold': False, 'border': 2, 'border_color': '#d9d9d9'})
    formats['fmt_wrap'].set_align('center')
    formats['fmt_wrap'].set_align('vcenter')
    formats['fmt_wrap'].set_text_wrap()


    formats['fmt_comment'] = wb.add_format({'font_size': 11, 'bold': False, 'border': 2, 'border_color': '#d9d9d9', 'text_wrap': True})
    formats['fmt_comment'].set_text_wrap()
    formats['fmt_comment'].set_align('vcenter')
    formats['fmt_comment'].set_align('left')

    formats['bold'] = wb.add_format({'font_size': 11, 'bold': True, 'text_wrap': True})
    formats['bold'].set_text_wrap()
    formats['bold'].set_align('vcenter')

    wrap = wb.add_format({'text_wrap': True, 'align': 'vcenter'})
    # set column width

    # ignore errors
    ws.ignore_errors({'number_stored_as_text': 'A1:V200'})


    width = [17, 16, 19.5, 10] + [10] * n_family + [12, 12, 12, 8, 9.5, 8, 13]
    for n, v in enumerate(width):
        ws.set_column(n, n, v)

    ws.set_column(11 + n_family, 11 + n_family, 90, cell_format=wrap)

    # first row
    row = 0
    ws.merge_range(row, 0, row, total_col, prj, cell_format=header1)

    note = 'Research Variants (need Sanger confirmation unless indicated under Comments)'
    row += 1
    ws.merge_range(row, 0, row, total_col, note, cell_format=header2)

    # note = 'new row'
    # row += 1
    # ws.merge_range(row, 0, row, total_col, note, cell_format=header2)

    # sort the rank with in he same sv type
    col_comment = idx['phenotype']
    # first sort by comment len
    # v is a list
    sv_selected_sorted = [(k, sorted(v, key=lambda _: len(_[col_comment]), reverse=True)) for k, v in sv_selected.items()]

    # then sort by gn rank
    # sv_selected_sorted = [(k, sorted(v, key=lambda _: int(_[0]))) for k, v in sv_selected.items()]
    # print([v[0][:col_comment] for k, v in sv_selected_sorted])
    sv_selected_sorted = [(k, sorted(v, key=lambda _: int(_[0]))) for k, v in sv_selected_sorted]

    # sort the order of sv type, denomo on top or hetero on top
    sv_selected_sorted = sorted(sv_selected_sorted, key=lambda _: int(_[1][0][0]))

    row += 1

    for sv_type_long, info in sv_selected_sorted:
        # sv_type_long = sv_type_conversion[sv_type]
        row = add_header(ws, row, sv_type_long, family_info, formats)
        prev_gn = 'demo111'
        gn_record = {}
        for info1 in info:
            gn = info1[idx['gn']]
            if gn == prev_gn:
                dup_gn = 1
            else:
                dup_gn = None
                gn_record[gn] = {'row_start': row}

            # print(gn, dup_gn)
            prev_gn = gn
            res, comment_info = add_data(ws, row, info1, idx, n_family, formats, sv_caller=sv_caller, dup_gn=dup_gn)
            if res is not None:
                row = res
                gn_record[gn]['row_end'] = row - 1
            if comment_info is not None:
                gn_record[gn]['comment_info'] = comment_info


        for gn, v in gn_record.items():
            row_start = v['row_start']
            row_end = v['row_end']
            comment_info = v['comment_info']

            col_gn = 0
            ws.merge_range(row_start, col_gn, row_end, col_gn, gn, cell_format=formats['fmt_txt'])


            height = comment_info['height']
            comment_list = comment_info['comment_list']
            fmt_types = set(comment_list[::2])

            # report xlsx,
            col_comment = 11 + n_family
            ws.merge_range(row_start, col_comment, row_end, col_comment, '', cell_format=formats['fmt_comment'])

            # if you need to use write_rich_string, the format types must be more than 1
            if len(fmt_types) == 1:
            # print('writing single format cell')
                comment_list = comment_list[1::2]
                fmt = list(fmt_types)[0]
                ws.write_string(row_start, col_comment, ' '.join(comment_list), fmt)
            else:
                res = ws.write_rich_string(row_start, col_comment, *comment_list)
            if res:
                print(f'fail to write rich text: {gn}')
                # print(comment_list)
                # print(len(comment_list))

            ws.set_row(row_start + 3, height * 14)

    wb.close()



def add_header(ws, row, sv_type_long, family_info, formats):
    n_family = len(family_info)
    # sv_type line
    ws.merge_range(row, 0, row, 11 + n_family, f'Structural Variants({sv_type_long})', cell_format=formats['header_sv'])

    row += 1
    merged_cell_col_idx = [0, 2, 3] + [_ + 4 for _ in range(n_family)] + [_ + 7 + n_family for _ in [0, 1, 2, 3, 4]]
    merged_cell_value = ['Gene', 'Change', 'Effect'] + family_info +['Baylor WGS', 'Emedgene', 'Yu Shyr', 'PreUDN Panel', 'Comments']
    for col, v in zip(merged_cell_col_idx, merged_cell_value):
        ws.merge_range(row, col, row + 3, col, v, cell_format=formats['header_main'])


    # pos
    col_idx = [1] + [4 + n_family + _ for _ in [0, 1, 2]]
    values = [['Chr', 'Position', 'sv_len', ''], ['Quality', 'GQ', 'Coverage', ''], ['dup', 'dup_source', 'del', 'del_source'], ['Missense Z', 'LoF pLI', '', '']]

    for col, v in zip(col_idx, values):
        for i_row, i in enumerate(v):
            ws.write_string(row + i_row, col, i, cell_format=formats['header_main'])
    return row + 4


def add_data(ws, row, data, idx, n_family, formats, sv_caller='dragen', dup_gn=False):
    try:
        # ["mut_type", "AMELIE", 'filter', "anno_id", "chr_", "pos_s", "pos_e", "sv_type", "qual", "exon_span_tag", "gn", "sv_len", "exon_span", "gain_af_max", "loss_af_max", "ddd_mode", "ddd_disease", "omim", "phenotype", "inheritance", "annot_ranking"]
        rank, chr_, s, e, sv_type, qual, exon_span_tag, gn, sv_len, exon_span, gain_afmax, loss_afmax, gain_source, loss_source, phenotype = [data[idx[_]] for _ in ['AMELIE', 'chr_', 'pos_s', 'pos_e', 'sv_type', 'qual', 'exon_span_tag', 'gn', 'sv_len', 'exon_span', 'gain_af_max', 'loss_af_max', 'gain_source', 'loss_source', 'phenotype']]
    except:
        print(f'wrong line format!, columns less than 19: {len(data)} ')
        print(data[:19])
    # exon_span_tag = '' if exon_span_tag == np.nan else exon_span_tag

    col_in_proband = idx['annot_ranking'] + 1
    cn_proband = data[col_in_proband].replace('@', '\n')


    if sv_caller == 'dragen':
        cn_family = [_.replace('@', '\n') for _ in list(data[col_in_proband + 1:])]
    else:
        cn_family = [_ for _ in list(data[col_in_proband + 1:])]

    if len(cn_family) != n_family - 1:
        print(f'error, the family number count in file({len(cn_family)}) and Family info({n_family}) list are differnet !')
        sys.exit(1)

    pos = f'{s}-{e}'

    comment = phenotype

    # write the pos
    col = 1
    try:
        ws.write_string(row, col, chr_, cell_format=formats['fmt_txt'])
    except:
        print('wrong row, data = {data}, chr={chr_}')
        return None
    ws.merge_range(row + 1, col, row + 2, col, pos, cell_format=formats['fmt_txt'])
    ws.write_string(row + 3, col, sv_len, cell_format=formats['fmt_txt'])

    # write merged cells


    # set to dump None, before the function is running properly
    # dup_gn = None

    # the merge of gn and comment should be done outside of the func
    col_idx = [3] + [4 + _ for _ in range(n_family)] + [4 + n_family + _ for _ in [3, 4, 6]]
    values = [sv_type, cn_proband] + cn_family + ['', '', '']
    for col, v in zip(col_idx, values):
        ws.merge_range(row, col, row + 3, col, v, cell_format=formats['fmt_wrap'])

    # if dup_gn:
    #     col_idx = [3] + [4 + _ for _ in range(n_family)] + [4 + n_family + _ for _ in [3, 4, 6]]
    #     values = [sv_type, cn_proband] + cn_family + ['', '', '']
    #     for col, v in zip(col_idx, values):
    #         ws.merge_range(row, col, row + 3, col, v, cell_format=formats['fmt_txt'])
    #     col_gn = 0
    #     row_prev_gn = dup_gn['row']
    #     ws.merge_range(row_prev_gn, col_gn, row + 3, col_gn, gn)
    # else:
    #     col_idx = [0, 3] + [4 + _ for _ in range(n_family)] + [4 + n_family + _ for _ in [3, 4, 6]]
    #     values = [gn, sv_type, cn_proband] + cn_family + ['', '', '']
    #     for col, v in zip(col_idx, values):
    #         ws.merge_range(row, col, row + 3, col, v, cell_format=formats['fmt_txt'])



    # write the split cols
    col_idx = [2] + [4 + n_family + _ for _ in [0, 1, 2, 5]]
    values = [
            ['', exon_span, exon_span_tag, ''],
            [qual, '', '', ''],
            [gain_afmax, gain_source, loss_afmax, loss_source],  #['dup', 'dup_source', 'del', 'del_source']
            ['', '', '', ''],
            [rank, '✔', '', '']
    ]
    for col, v in zip(col_idx, values):
       for n, iv in enumerate(v):
           iv = str(iv)
           ws.write_string(row + n, col, iv, cell_format=formats['fmt_txt'])

    # write the comment

    if dup_gn is None:
        bold = formats['bold']
        cell_format = formats['fmt_comment']


        bold_pattern = r'\*\*([\S][^*]+)\*\*'
        spans = []
        for i in re.finditer(bold_pattern, comment):
            spans.append(i.span())

        comment_list = []
        prev_plain_text_start = 0
        for s, e in spans:
            txt = comment[prev_plain_text_start: s]
            if txt:
                comment_list.extend([cell_format, txt])
            txt = comment[s+2: e-2]
            if txt.strip():
                comment_list.extend([bold, txt])
            prev_plain_text_start = e
        txt = comment[prev_plain_text_start: ]
        if txt:
            comment_list.extend([cell_format, txt])

        # print(comment_list)
        # ws.write(row, col, comment)

        # prepare to set the row height
        # each line is about 100 char, height = 15 px
        tmp = comment.split('\n')
        tmp = [int(len(_)/105) + 1 for _ in tmp]
        height = sum(tmp) - 3 # exclude the 3 merged row height
        # print(f"expected_lines = {height}")
        comment_info = {'comment_list': comment_list, 'height': height}
    else:
        comment_info = None

    return row + 4, comment_info



if __name__ == "__main__":
    import argparse as arg
    from argparse import RawTextHelpFormatter
    ps = arg.ArgumentParser(description=__doc__, formatter_class=RawTextHelpFormatter)
    ps.add_argument('prj', help="""case name/title""")
    ps.add_argument('fn', help="""selected SV list, 1st column = SV type(denovo/heter/xlink)
    2nd column is the rank
    last 2/3 column (col 21-end) must be the copy number of proband and family members
    """, nargs='?')
    # ps.add_argument('-family', help="""family info, put in the header line. sep by comma
    # like  'Proand 123456', 'Mother (unaff) 22222', 'Father(unaff) 33333'""", nargs='+')
    args = ps.parse_args()

    fn = args.fn
    prj = args.prj
    # family_info = ' '.join(args.family)
    # family_info = family_info.split(',')
    # family_info = [_.strip() for _ in family_info if _.strip()]
    pw = os.getcwd()
    if fn and not os.path.exists(fn):
        print(f'error, selected SV list file not exist ! {fn}')
        sys.exit(1)

    main(prj, pw, fn)
