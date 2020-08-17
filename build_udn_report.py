import sys, os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle, Color
import openpyxl
from openpyxl.utils import get_column_letter


# def s():
#     fno = 'test.xlsx'
#     wb.save(fno)
#     !open {fno}
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Original'

def main(prj, fn_selected_genes, family_info):

    """
    prj = UDNxxxx
    family_info = ['Proband UDN11111', 'Father UDN22222', 'Mother UDN33333']

    fn_selected_genes ,  first column should be the sv_type, valid = 'denovo, heter, xlink'
    second col, previous = amelie, now = rank, from 1 to 10
    last 2/3 column (col 21-end) must be the copy number of proband and family members


    # family info, is a list or tuple, like
    # ['Proand 123456', 'Mother (unaff) 22222', 'Father(unaff) 33333']
    """

    # group the selected SV into groups
    sv_selected = {}
    sv_type_conversion = {'denovo': 'De Novo',
                        'heter': 'Heterozygous',
                        'xlink': 'X-linked',
                        }

    # style / format
    bd = Side(style='medium', color='d9d9d9')
    colr_blue = Color(rgb='00dce6f1')  # light blue
    colr_gray = Color(rgb='00d9d9d9')  # light blue

    fmt_header = NamedStyle(name='header',
                            font=Font(size=11, bold=True,),
                            border=Border(left=bd, right=bd, top=bd, bottom=bd),
                            fill=PatternFill(fill_type='solid', fgColor=colr_blue),
                            alignment=Alignment(vertical='center', horizontal='center', wrap_text=True),
                            )

    fmt_sv_type = NamedStyle(name='sv_type',
                            font=Font(size=14, bold=True,),
                            fill=PatternFill(fill_type='solid', fgColor=colr_gray),
                            )

    fmt_data = NamedStyle(name='data',
                        border=Border(left=bd, right=bd, top=bd, bottom=bd),
                        alignment=Alignment(vertical='center', horizontal='center'),
                        )

    fmt_data_wrap = NamedStyle(name='data_wrap',
                        border=Border(left=bd, right=bd, top=bd, bottom=bd),
                        alignment=Alignment(vertical='center', horizontal='center', wrap_text=True),
                        )


    wb.add_named_style(fmt_data)
    wb.add_named_style(fmt_data_wrap)
    wb.add_named_style(fmt_header)
    wb.add_named_style(fmt_sv_type)

    data = [_.strip().split('\t') for _ in open(fn_selected_genes)]
    for i in data:
        sv_type = i[0]
        if sv_type not in sv_type_conversion:
            if i[1] != "AMELIE":
                print(f'wrong SV catergory found! {sv_type}')
            continue
        try:
            sv_selected[sv_type].append(i[1:])
        except:
            sv_selected[sv_type] = [i[1:]]
    # set the width
    set_column_width()

    # row1
    row = 1
    ws.merge_cells('A1:O1')
    ws['A1'] = prj
    ws['A1'].font = Font(sz=16, bold=True)

    note = 'Research Variants (need Sanger confirmation unless indicated under Comments)'
    row = 2
    ws.merge_cells('A2:O2')
    ws['A2'] = note
    ws['A2'].font = Font(sz=14, bold=True)

    # add real data
    row = 3
    for sv_type, v in sv_selected.items():
        sv_type_long = sv_type_conversion[sv_type]
        row = add_header(row, sv_type_long, family_info)
        for i_cnv in v:
            row = add_data(row, i_cnv, family_info)


def add_data(row, data, family_info):
    rank, _, chr_, s, e, sv_type, qual, exon_span_tag, gn, sv_len, exon_span, dgv_gain, dgv_loss, gnomad, _, ddd_disease, _, phenotype, inher = data[:19]
    cn_proband = data[20]
    cn_family = data[21:]
    if len(cn_family) != len(family_info) - 1:
        print(f'error, the family number count in file({len(cn_family)}) and Family info({len(family_info)}) list are differnet !')
        sys.exit(1)

    pos = f'{s}-{e}'
    dgv = f'{dgv_gain}/{dgv_loss}'
    comment = '\n'.join([f'{phenotype} {inher}', ddd_disease, ])

    # build the header
    col_raw = 0
    # gene
    col_raw += 1
    col = get_column_letter(col_raw)
    ws.merge_cells(f'{col}{row}:{col}{row+3}')
    ws[f'{col}{row}'].value = gn

    # pos
    col_raw += 1
    col = get_column_letter(col_raw)

    ws[f'{col}{row}'].value = chr_

    # pos
    ws.merge_cells(f'{col}{row+1}:{col}{row+2}')
    ws[f'{col}{row+1}'].value = pos


    # overlap
    ws[f'{col}{row+3}'].value = sv_len

    # change
    col_raw += 1
    col = get_column_letter(col_raw)

    ws[f'{col}{row+1}'].value = exon_span
    ws[f'{col}{row+2}'].value = exon_span_tag

    # effect
    col_raw += 1
    col = get_column_letter(col_raw)
    ws.merge_cells(f'{col}{row}:{col}{row+3}')
    ws[f'{col}{row}'].value = sv_type

    # proband
    col_raw += 1
    col = get_column_letter(col_raw)
    ws.merge_cells(f'{col}{row}:{col}{row+3}')
    ws[f'{col}{row}'].value = cn_proband

    # family member
    for i_cn in cn_family:
        col_raw += 1
        col = get_column_letter(col_raw)
        ws.merge_cells(f'{col}{row}:{col}{row+3}')
        ws[f'{col}{row}'].value = i_cn

    # quality
    col_raw += 1
    col = get_column_letter(col_raw)
    ws[f'{col}{row}'].value = qual

    # allele freq
    col_raw += 1
    col = get_column_letter(col_raw)
    ws[f'{col}{row}'].value = gnomad
    ws[f'{col}{row+1}'].value = dgv

    # scores-2
    col_raw += 1

    # baylor
    col_raw += 1
    col = get_column_letter(col_raw)
    ws.merge_cells(f'{col}{row+1}:{col}{row+3}')

    # emedgene
    col_raw += 1
    col = get_column_letter(col_raw)
    ws.merge_cells(f'{col}{row}:{col}{row+3}')

    # yushyr
    col_raw += 1
    col = get_column_letter(col_raw)
    ws[f'{col}{row}'].value = rank
    ws[f'{col}{row + 1}'].value = '✔'

    # panel
    col_raw += 1
    col = get_column_letter(col_raw)
    ws.merge_cells(f'{col}{row}:{col}{row+3}')


    # comments
    col_raw += 1
    col = get_column_letter(col_raw)
    ws.merge_cells(f'{col}{row}:{col}{row+3}')
    ws[f'{col}{row}'].value = comment

    wrap_col = set([2, 15])
    # apply the format
    for irow in ws[f'A{row}:O{row+3}']:
        for cell in irow:
            if cell.column in wrap_col:
                cell.style = 'data_wrap'
            else:
                cell.style = 'data'
    return row + 4


def set_column_width(width=None):
    """
    specify the width, should be a list with 15 elements
    """
    width = width or [17, 16, 17, 12, 10, 10, 10, 12, 12, 12, 8, 9.5, 8, 13, 95]
    assert len(width) == 15, 'Set column width, the element number must be 15'
    for n, v in enumerate(width):
        ws.column_dimensions[get_column_letter(n+1)].width = v


# add_header(5, 'De Novo', ['proband', 'Father', 'Mother'])
def add_header(row, sv_type_long, family_info):
    """
    row = new row number
    family_info is a list/tuple, stores the sample name for the header, e.g.
    ['Proand 123456', 'Mother (unaff) 22222', 'Father(unaff) 33333']
    """

    # sv_type line
    ws.merge_cells(f'A{row}:O{row}')
    cell = f'A{row}'
    ws[cell].value = f'Structural Variants({sv_type_long})'
    ws[cell].style = "sv_type"

    # build the header
    col_raw = 0
    # gene
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'Gene'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    ws[f'{col}{row+1}'].value = v

    # pos
    col_raw += 1
    col = get_column_letter(col_raw)
    v_cells = ['Chr', 'Position', 'rs#', '']
    for cell, v in zip(ws[f'{col}{row+1}:{col}{row+4}'], v_cells):
        cell = cell[0]
        cell.value = v

    # change
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'Change'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    cell = f'{col}{row+1}'
    ws[cell].value = v

    # effect
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'Effect'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    cell = f'{col}{row+1}'
    ws[cell].value = v

    # family member
    for isample in family_info:
        col_raw += 1
        col = get_column_letter(col_raw)
        ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
        cell = f'{col}{row+1}'
        ws[cell].value = isample

    # quality
    col_raw += 1
    col = get_column_letter(col_raw)
    v_cells = ['Quality', 'GQ', 'Coverage', '']
    for cell, v in zip(ws[f'{col}{row+1}:{col}{row+4}'], v_cells):
        cell = cell[0]
        cell.value = v

    # scores
    col_raw += 1
    col = get_column_letter(col_raw)
    v_cells = ['GnomAD', 'DGV(dup/del)', 'GERP', 'CADD']
    for cell, v in zip(ws[f'{col}{row+1}:{col}{row+4}'], v_cells):
        cell = cell[0]
        cell.value = v

    # scores-2
    col_raw += 1
    col = get_column_letter(col_raw)
    v_cells = ['Missense Z', 'LoF pLI', '', '']
    for cell, v in zip(ws[f'{col}{row+1}:{col}{row+4}'], v_cells):
        cell = cell[0]
        cell.value = v

    # baylor
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'Baylor WGS'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    cell = f'{col}{row+1}'
    ws[cell].value = v

    # emedgene
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'Emedgene'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    cell = f'{col}{row+1}'
    ws[cell].value = v

    # yushyr
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'Yu Shyr'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    cell = f'{col}{row+1}'
    ws[cell].value = v

    # panel
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'PreUDN Panel'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    cell = f'{col}{row+1}'
    ws[cell].value = v

    # comments
    col_raw += 1
    col = get_column_letter(col_raw)
    v = 'Comments'
    ws.merge_cells(f'{col}{row+1}:{col}{row+4}')
    cell = f'{col}{row+1}'
    ws[cell].value = v

    # apply the format
    for irow in ws[f'A{row+1}:O{row+4}']:
        for cell in irow:
            cell.style = 'header'
    return row + 5

if __name__ == "__main__":
    import argparse as arg
    from argparse import RawTextHelpFormatter
    ps = arg.ArgumentParser(description=__doc__, formatter_class=RawTextHelpFormatter)
    ps.add_argument('prj', help="""case name/title""")
    ps.add_argument('fn', help="""selected SV list, 1st column = SV type(denovo/heter/xlink)
    2nd column is the rank
    last 2/3 column (col 21-end) must be the copy number of proband and family members
    """)
    ps.add_argument('family', help="""family info, put in the header line. sep by comma
    like  'Proand 123456', 'Mother (unaff) 22222', 'Father(unaff) 33333'""", nargs='+')
    args = ps.parse_args()

    fn = args.fn
    prj = args.prj
    family_info = ' '.join(args.family)
    family_info = family_info.split(',')
    family_info = [_.strip() for _ in family_info if _.strip()]

    if not os.path.exists(fn):
        print(f'error, selected SV list file not exist ! {fn}')
        sys.exit(1)

    main(prj, fn, family_info)
    fnout = f'{prj}.report.xlsx'
    wb.save(fnout)
    os.system(f'open {fnout}')
